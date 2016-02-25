import csv 
import sys
import string
import indicoio
from TwitterSearch import *
from indicoio import political, sentiment, language, text_tags, keywords, fer, facial_features, image_features
<<<<<<< HEAD
<<<<<<< Updated upstream
=======
from openpyxl import Workbook
from openpyxl import load_workbook
>>>>>>> origin/master

def remove_http(source_string, replace_what): #defining a funciton for pulling out the links in the tweets, that rhymes haha
    head, sep, tail = source_string.rpartition(replace_what)
    return head 

def convert(schoolName,key1,key2,key3,key4):
    #importing the databases of campuses & locations

    campus_loc = {} #{name:[long:lat]}
    with open('campuses.csv') as f:
        reader = csv.reader(f)
        reader.next()
        for row in reader:
            campus_loc[row[1]] = [float(row[3]), float(row[4])]

    school = schoolName #takes input and renames it for convenience's sake since every other variable name is schoolName #whoops

    """
    counter = 0 #this entire block will never be run unless we take user input
    while school not in campus_loc.keys():
            counter += 1
            print "Sorry! Doesn't look like we have that school. Try again?"
            if counter > 5:
                    print "Wow, you really suck at this."
            school = raw_input("> ")
    """
    location = campus_loc[school] #gets the location from campus_loc using the school name

    ############## begin Twitter API call now

    try:
        tso = TwitterSearchOrder() # create a TwitterSearchOrder object
        tso.set_keywords([' ']) # let's define all words we would like to have a look for
        tso.set_geocode(location[1],location[0],1)
        #tso.set_language('de') # we want to see German tweets only
        tso.set_include_entities(False) # and don't give us all those entity information

        # it's about time to create a TwitterSearch object with our secret tokens
        ts = TwitterSearch(
            consumer_key = str(key1),
            consumer_secret = str(key2),
            access_token = str(key3),
            access_token_secret = str(key4)
            #consumer_key = 'ZrOjYYBF9ALwFlHjnJ6uRwBk6',
            #consumer_secret = 'QraDsU5pdeSTT7qNcReCJZYwmX94Q6S7yb0EcagPumCNmafufq',
            #access_token = '16986893-nUTUdtEcf2HffYJYpVJYbR2p85EeSmvvZSD2VsCLS',
            #access_token_secret = 'I8V9DJNFWQX0wh0gNbfLT2bAWpj4uS7rtUioFqyDjnVf3'
           # consumer_key = 'naPGQJt3L75sZiEZAplsETEp1',
        #   consumer_secret = 'rqyRlly80nw1XUGxt7ySR2ZUvQiqDKpbj8kEgDXJ63U1AWbvAr',
           # access_token = '16986893-Zq2L75kDTGR0l3AnsSN1ZXkzw4o8NuBssLCyvxkES',
           # access_token_secret = 'woWPf0kgYoJcz6w4mNUBc8tt2bWqwdSCoMPEYXsWU9Y6w'
         )

         # this is where the fun actually starts :)
        collegeTweets = []
        counter = 0
        for tweet in ts.search_tweets_iterable(tso):
            if counter < 50:
                collegeTweets.append(tweet['text'].encode('ascii','ignore')) #ignores any ASCII
                collegeTweets[-1] = string.replace(collegeTweets[-1],"#","") #pulls out pound signs to make hashtags part of the sentence/string
                collegeTweets[-1] = string.replace(collegeTweets[-1],"\n","")  #pulls out any returns which fuck up the sentence
                collegeTweets[-1] = string.replace(collegeTweets[-1],"@","") #pulls out any mentions
                collegeTweets[-1] = remove_http(collegeTweets[-1],"http") #pulls out any links (aka to pictures since the majority of these are linked from Instagram)
                counter += 1
                collegeTweets[-1] += "a"

    except TwitterSearchException as e:
        print(e)    

    ############### begin Indico API call now

    if not collegeTweets:
        return 0.5

    indicoio.config.api_key = "f09f509655f721e3adac6df5b35abfed"
    api_key_Lisa = "f09f509655f721e3adac6df5b35abfed"

    sentementCollegeTweets = sentiment(collegeTweets) #take the Twitter string agnd put it into our own string because we needed an array of the indico results

    average = 0.0
    for i in sentementCollegeTweets:
        average += i
    average = average/len(sentementCollegeTweets)

    print average



for z in range(10):
    print "Enter a college name"
    school1 = raw_input("> ")
    convert(school1, 'ArKJR0f0L0k82OBcuixtzrpco','bD0B7IHKlmXiGoZIZRENXzGllguLcLAIPaB82kyssmehHeqZtP','3263098490-wANsX5QThAqOd75QTtFPYyV9qWiRCn6503MzsdN','hF3R5UqPUoTzjV1jcTfYa3peKckdxMClj8gfgdhX5YT33')
for w in range(10):
    print "Enter a college name"
    school1 = raw_input("> ")
    convert(school1,'yAmfVcbhF9a9couq4NYdkbKsb','59FF2r0StRym1QGF97ILIIwqu7okVwk2Wmhfx38o7DxOxuHqEp','3263098490-ZOgNiqlRduhttM5WSgXMvLT37b2g6gGrA2OVx8o','MByjTrBOSRs31S123infoy0coY2cPzm5xT00bTJIJB2KH')
    



<<<<<<< HEAD
=======
<<<<<<< HEAD

def remove_http(source_string, replace_what): #defining a funciton for pulling out the links in the tweets, that rhymes haha
    head, sep, tail = source_string.rpartition(replace_what)
    return head 

def convert(school):
    #importing the databases of campuses & locations
    campus_loc = {} #{name:[long:lat]}
    with open('campuses.csv') as f:
        reader = csv.reader(f)
        reader.next()
        for row in reader:
            campus_loc[row[1]] = [float(row[3]), float(row[4])]

    print "What school would you like to look at?"
    school = raw_input("> ") 
    print "We're going to look at %s." %school 

    counter = 0 #this entire block will never be run unless we take user input
    while school not in campus_loc.keys():
            counter += 1
            print "Sorry! Doesn't look like we have that school. Try again?"
            if counter > 5:
                    print "Wow, you really suck at this."
            school = raw_input("> ")

    location = campus_loc[school] #gets the location from campus_loc using the school name

    ############## begin Twitter API call now

    try:
        tso = TwitterSearchOrder() # create a TwitterSearchOrder object
        tso.set_keywords([' ']) # let's define all words we would like to have a look for
        tso.set_geocode(location[1],location[0],1)
        #tso.set_language('de') # we want to see German tweets only
        tso.set_include_entities(False) # and don't give us all those entity information

        # it's about time to create a TwitterSearch object with our secret tokens
        ts = TwitterSearch(
            consumer_key = 'naPGQJt3L75sZiEZAplsETEp1',
            consumer_secret = 'rqyRlly80nw1XUGxt7ySR2ZUvQiqDKpbj8kEgDXJ63U1AWbvAr',
            access_token = '16986893-Zq2L75kDTGR0l3AnsSN1ZXkzw4o8NuBssLCyvxkES',
            access_token_secret = 'woWPf0kgYoJcz6w4mNUBc8tt2bWqwdSCoMPEYXsWU9Y6w'
         )

         # this is where the fun actually starts :)
        collegeTweets = []
        counter = 0
        for tweet in ts.search_tweets_iterable(tso):
            if counter < 321:
                collegeTweets.append(tweet['text'].encode('ascii','ignore')) #ignores any ASCII
                collegeTweets[-1] = string.replace(collegeTweets[-1],"#","") #pulls out pound signs to make hashtags part of the sentence/string
                collegeTweets[-1] = string.replace(collegeTweets[-1],"\n","")  #pulls out any returns which fuck up the sentence
                collegeTweets[-1] = string.replace(collegeTweets[-1],"@","") #pulls out any mentions
                collegeTweets[-1] = remove_http(collegeTweets[-1],"http") #pulls out any links (aka to pictures since the majority of these are linked from Instagram)
                counter += 1
                collegeTweets[-1] += "a"
                print collegeTweets[-1]

    except TwitterSearchException as e:
        print(e)    

    ############### begin Indico API call now

    indicoio.config.api_key = "f09f509655f721e3adac6df5b35abfed"
    api_key_Lisa = "f09f509655f721e3adac6df5b35abfed"

    sentementCollegeTweets = sentiment(collegeTweets) #take the Twitter string agnd put it into our own string because we needed an array of the indico results

    average = 0.0
    for i in sentementCollegeTweets:
        average += i
    average = average/len(sentementCollegeTweets)

    print average
    return average



wb = Workbook()
ws = wb.active
school1 = convert("Pepperdine University")

ws['A1'] = school1
ws.append([1, 2, 3])

wb.save("sample.xlsx")



<<<<<<< HEAD
#importing the databases of campuses & locations
campus_loc = {} #{name:[long:lat]}
with open('campuses.csv') as f:
    reader = csv.reader(f)
    reader.next()
    for row in reader:
        campus_loc[row[1]] = [float(row[3]), float(row[4])]

school = "Harvard University"
#school = raw_input("> ") - what you implement if you want to ask the user for a school name
print "We're going to look at %s." %school 

counter = 0 #this entire block will never be run unless we take user input
while school not in campus_loc.keys():
	counter += 1
	print "Sorry! Doesn't look like we have that school. Try again?"
	if counter > 5:
		print "Wow, you really suck at this."
	school = raw_input("> ")

location = campus_loc[school] #gets the location from campus_loc using the school name

############## begin Twitter API call now

try:
    tso = TwitterSearchOrder() # create a TwitterSearchOrder object
    tso.set_keywords([' ']) # let's define all words we would like to have a look for
    tso.set_geocode(location[1],location[0],50)
    #tso.set_language('de') # we want to see German tweets only
    tso.set_include_entities(False) # and don't give us all those entity information

    # it's about time to create a TwitterSearch object with our secret tokens
    ts = TwitterSearch(
        consumer_key = 'naPGQJt3L75sZiEZAplsETEp1',
        consumer_secret = 'rqyRlly80nw1XUGxt7ySR2ZUvQiqDKpbj8kEgDXJ63U1AWbvAr',
        access_token = '16986893-Zq2L75kDTGR0l3AnsSN1ZXkzw4o8NuBssLCyvxkES',
        access_token_secret = 'woWPf0kgYoJcz6w4mNUBc8tt2bWqwdSCoMPEYXsWU9Y6w'
     )
    tweets = []
     # this is where the fun actually starts :)
    for tweet in ts.search_tweets_iterable(tso):
        
        tweets.append((tweet['text']).encode('ascii','ignore') )
        print( '@%s tweeted: %s' % ( tweet['user']['screen_name'], tweet['text'] ) )

except TwitterSearchException as e: # take care of all those ugly errors if there are some
    print(e)
=======
>>>>>>> origin/master
=======

def remove_http(source_string, replace_what): #defining a funciton for pulling out the links in the tweets, that rhymes haha
    head, sep, tail = source_string.rpartition(replace_what)
    return head 

def convert(school):
    #importing the databases of campuses & locations
    campus_loc = {} #{name:[long:lat]}
    with open('campuses.csv') as f:
        reader = csv.reader(f)
        reader.next()
        for row in reader:
            campus_loc[row[1]] = [float(row[3]), float(row[4])]

    print "What school would you like to look at?"
    school = raw_input("> ") 
    print "We're going to look at %s." %school 

    counter = 0 #this entire block will never be run unless we take user input
    while school not in campus_loc.keys():
            counter += 1
            print "Sorry! Doesn't look like we have that school. Try again?"
            if counter > 5:
                    print "Wow, you really suck at this."
            school = raw_input("> ")

    location = campus_loc[school] #gets the location from campus_loc using the school name

    ############## begin Twitter API call now

    try:
        tso = TwitterSearchOrder() # create a TwitterSearchOrder object
        tso.set_keywords([' ']) # let's define all words we would like to have a look for
        tso.set_geocode(location[1],location[0],1)
        #tso.set_language('de') # we want to see German tweets only
        tso.set_include_entities(False) # and don't give us all those entity information

        # it's about time to create a TwitterSearch object with our secret tokens
        ts = TwitterSearch(
            consumer_key = 'naPGQJt3L75sZiEZAplsETEp1',
            consumer_secret = 'rqyRlly80nw1XUGxt7ySR2ZUvQiqDKpbj8kEgDXJ63U1AWbvAr',
            access_token = '16986893-Zq2L75kDTGR0l3AnsSN1ZXkzw4o8NuBssLCyvxkES',
            access_token_secret = 'woWPf0kgYoJcz6w4mNUBc8tt2bWqwdSCoMPEYXsWU9Y6w'
         )

         # this is where the fun actually starts :)
        collegeTweets = []
        counter = 0
        for tweet in ts.search_tweets_iterable(tso):
            if counter < 321:
                collegeTweets.append(tweet['text'].encode('ascii','ignore')) #ignores any ASCII
                collegeTweets[-1] = string.replace(collegeTweets[-1],"#","") #pulls out pound signs to make hashtags part of the sentence/string
                collegeTweets[-1] = string.replace(collegeTweets[-1],"\n","")  #pulls out any returns which fuck up the sentence
                collegeTweets[-1] = string.replace(collegeTweets[-1],"@","") #pulls out any mentions
                collegeTweets[-1] = remove_http(collegeTweets[-1],"http") #pulls out any links (aka to pictures since the majority of these are linked from Instagram)
                counter += 1
                collegeTweets[-1] += "a"
                print collegeTweets[-1]

    except TwitterSearchException as e:
        print(e)    

    ############### begin Indico API call now

    indicoio.config.api_key = "f09f509655f721e3adac6df5b35abfed"
    api_key_Lisa = "f09f509655f721e3adac6df5b35abfed"

    sentementCollegeTweets = sentiment(collegeTweets) #take the Twitter string agnd put it into our own string because we needed an array of the indico results

    average = 0.0
    for i in sentementCollegeTweets:
        average += i
    average = average/len(sentementCollegeTweets)

    print average
    return average



wb = Workbook()
ws = wb.active
school1 = convert("Pepperdine University")

ws['A1'] = school1
ws.append([1, 2, 3])

wb.save("sample.xlsx")



>>>>>>> origin/master
>>>>>>> Stashed changes
=======
print campus_final
>>>>>>> origin/master

#wb1.save("sample.xlsx")
